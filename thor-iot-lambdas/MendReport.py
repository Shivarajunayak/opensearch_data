import json
import openpyxl
from openpyxl import Workbook,load_workbook,drawing
from openpyxl.styles import Alignment,PatternFill,Font,Side,Border
import os
import pandas as pd
import sys

def start_creating_workbook(*args):


        # Creating the Project and Repository name from teh arguments
        pname = ""
        rname = ""

        for item in args[1].split("_"):
                pname = pname + item + " "

        for item in args[2].split("_"):
                rname = rname + item + " "
                

        # Loading the existing JSON file
        f = open(args[0],'r')
        data = json.load(f)

        try:

                # Define border styles
                thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))

                # Working on adding the entries to the dataframes present in the sheet
                firstdataframe = pd.DataFrame(columns=['Library','Licenses'])
                for libraries in data["libraries"]:

                        first_entry = f"{libraries['name']} - {libraries['version']}"
                        if libraries['licenses']:
                                if len(libraries['licenses']) > 1:
                                                second_entry = ''
                                                for ls in range(len(libraries['licenses'])):
                                                        second_entry+=f"{libraries['licenses'][ls]['name']}, "
                                                second_entry = second_entry[:-2]

                                else:
                                                second_entry = f"{libraries['licenses'][0]['name']}"
                        else:
                                second_entry = "NULL"

                        # Adding the entry to the dataframe 
                        firstdataframe.loc[len(firstdataframe.index)] = [first_entry, second_entry]

                firstdataframe.sort_values(by=['Library'],inplace=True,ignore_index=True)


                # Initializing the new excel file
                wb = Workbook()
                wb.remove(wb.active)
                ws = wb.create_sheet("Inventory",index=2)
                
                # Adding Project, repository name and build-id of the project
                ws["A1"] = f"Project Name : {pname}"
                ws["A1"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A1"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A1"].font = Font(bold=True)
                ws["A1"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                
                ws["A3"] = f"Repository Name : {rname}"
                ws["A3"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A3"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A3"].font = Font(bold=True)
                ws["A3"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                
                ws["A5"] = f"Build ID : {args[3]}\n (Click here for Mend Dashboard)"
                ws["A5"].style = "Hyperlink"
                ws["A5"].hyperlink = f'{args[4]}'
                ws["A5"].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                ws["A5"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A5"].font = Font(bold=True,color='000000')
                ws["A5"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws.merge_cells(start_row = 1,start_column=1,end_row = 2,end_column=2)
                ws.merge_cells(start_row = 3,start_column=1,end_row = 4,end_column=2)
                ws.merge_cells(start_row = 5,start_column=1,end_row = 6,end_column=2)
                ws.freeze_panes = ws["C7"]

                # Creating the table headers (for 'Library' and 'Licenses') in the 'Inventory' sheet
                ws["A8"] = "Library"
                ws["A8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["A8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A8"].font = Font(bold=True)
                ws.column_dimensions['A'].width = 45
                ws["A8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["B8"] = "Licenses"
                ws["B8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["B8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["B8"].font = Font(bold=True)
                ws.column_dimensions['B'].width = 30
                ws["B8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                

                ws.merge_cells(start_row = 8,start_column=1,end_row = 9,end_column=1)
                ws.merge_cells(start_row = 8,start_column=2,end_row = 9,end_column=2)

                # Write the DataFrame data to the specified location in the worksheet
                for idx, row in firstdataframe.iterrows():
                        for col_idx, value in enumerate(row, start=1):
                                ws.cell(row=10 + idx, column=col_idx, value=value)
                                ws.cell(row=10 + idx, column=col_idx).border = thin_border
                                ws.cell(row=10 + idx, column=col_idx).alignment = Alignment(horizontal='center',vertical='center')

                                if col_idx == 1: 
                                                ws.cell(row=10 + idx, column=col_idx).alignment = Alignment(horizontal='left',vertical='center')
                                        
                ws.title = f'Inventory ({len(firstdataframe)})'

                """
                Started working on vulnerabilities sheet
                """

                # Working on adding the entries to the dataframes present in the sheet
                seconddataframe = pd.DataFrame(columns=['Severity','Vulnerabilty','Date','Library','Top Fix'])
                for libraries in data["libraries"]:
                        
                        if len(libraries["vulnerabilities"]) == 1:

                                first_entry = f"{libraries['vulnerabilities'][0]['severity']} : {libraries['vulnerabilities'][0]['score']}"
                                second_entry = f"{libraries['vulnerabilities'][0]['name']}"
                                third_entry = f"{libraries['vulnerabilities'][0]['publishDate']}"
                                fourth_entry = f"{libraries['name']}"

                                if "topFix" in libraries['vulnerabilities'][0].keys():
                                        fifth_entry = f"""{libraries['vulnerabilities'][0]['topFix']['fixResolution']}\n{libraries['vulnerabilities'][0]['topFix']['url']}"""
                                        
                                else:
                                        fifth_entry = "No Fix Information Available"

                                seconddataframe.loc[len(seconddataframe.index)] = [first_entry, second_entry, third_entry,fourth_entry, fifth_entry]
                        
                        elif len(libraries["vulnerabilities"]) > 1:

                                for vul in libraries["vulnerabilities"]:

                                        first_entry = f"{vul['severity']} : {vul['score']}"
                                        second_entry = f"{vul['name']}"
                                        third_entry = f"{vul['publishDate']}"
                                        fourth_entry = f"{libraries['name']}"

                                        if "topFix" in vul.keys():
                                                fifth_entry = f"""{vul['topFix']['fixResolution']}\n{vul['topFix']['url']}"""
                                        else:
                                                fifth_entry = "No Fix Information Available"

                                        seconddataframe.loc[len(seconddataframe.index)] = [first_entry, second_entry, third_entry,fourth_entry, fifth_entry]


                ws = wb.create_sheet(f'Security vulnerabilties ({len(seconddataframe)})',index=1)

                # Adding Project, repository name and build-id of the project
                ws["A1"] = f"Project Name : {pname}"
                ws["A1"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A1"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A1"].font = Font(bold=True)
                ws["A1"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                
                ws["A3"] = f"Repository Name : {rname}"
                ws["A3"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A3"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A3"].font = Font(bold=True)
                ws["A3"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                
                ws["A5"] = f"Build ID : {args[3]}\n (Click here for Mend Dashboard)"
                ws["A5"].style = "Hyperlink"
                ws["A5"].hyperlink = f'{args[4]}'
                ws["A5"].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                ws["A5"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A5"].font = Font(bold=True,color='000000')
                ws["A5"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws.merge_cells(start_row = 1,start_column=1,end_row = 2,end_column=5)
                ws.merge_cells(start_row = 3,start_column=1,end_row = 4,end_column=5)
                ws.merge_cells(start_row = 5,start_column=1,end_row = 6,end_column=5)
                ws.freeze_panes = ws["F7"]

                # Creating the table headers (for 'Library' and 'Licenses') in the 'Inventory' sheet
                ws["A8"] = 'Severity'
                ws["A8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["A8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A8"].font = Font(bold=True)
                ws.column_dimensions['A'].width = 20
                ws["A8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["B8"] = 'Vulnerabilty'
                ws["B8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["B8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["B8"].font = Font(bold=True)
                ws.column_dimensions['B'].width = 20
                ws["B8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["C8"] = 'Date'
                ws["C8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["C8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["C8"].font = Font(bold=True)
                ws.column_dimensions['C'].width = 20
                ws["C8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["D8"] = 'Library'
                ws["D8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["D8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["D8"].font = Font(bold=True)
                ws.column_dimensions['D'].width = 40
                ws["D8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["E8"] = 'Top Fix'
                ws["E8"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["E8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["E8"].font = Font(bold=True)
                ws.column_dimensions['E'].width = 85
                ws["E8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws.merge_cells(start_row = 8,start_column=1,end_row = 9,end_column=1)
                ws.merge_cells(start_row = 8,start_column=2,end_row = 9,end_column=2)
                ws.merge_cells(start_row = 8,start_column=3,end_row = 9,end_column=3)
                ws.merge_cells(start_row = 8,start_column=4,end_row = 9,end_column=4)
                ws.merge_cells(start_row = 8,start_column=5,end_row = 9,end_column=5)
                ws.merge_cells(start_row = 8,start_column=6,end_row = 9,end_column=6)
                ws.merge_cells(start_row = 8,start_column=7,end_row = 9,end_column=7)
                ws.merge_cells(start_row = 8,start_column=8,end_row = 9,end_column=8)


                if len(seconddataframe): 

                        # Write the DataFrame data to the specified location in the worksheet
                        for idx, row in seconddataframe.iterrows():
                                
                                ws.row_dimensions[10+idx].height = 50
                                for col_idx, value in enumerate(row, start=1):
                                                ws.cell(row=10 + idx, column=col_idx, value=value)
                                                ws.cell(row=10 + idx, column=col_idx).border = thin_border
                                                ws.cell(row=10 + idx, column=col_idx).alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')


                """
                Started working on License risks sheet
                """  

                # Working on adding the entries to the dataframes present in the sheet
                thirddataframe = pd.DataFrame(columns=['License','Risk','Occurrences'])

                # loop for storing the licenses names and their counts in dictionary
                lic_dic = {}
                lic_list = []
                for libraries in data["libraries"]:
                        
                        if libraries['licenses']:
                                if len(libraries['licenses']) > 1:
                                                
                                                for ls in range(len(libraries['licenses'])):
                                                        lic_list.append(libraries['licenses'][ls]['name'])  

                                else:
                                                lic_list.append(libraries['licenses'][0]['name'])

                for lic in lic_list:

                        if lic in lic_dic:
                                lic_dic[lic] += 1
                        else:
                                lic_dic.update({lic:1})
                
                lic_dic = dict(sorted(lic_dic.items()))
                
                # Default dictionary to store the license with copyrightscore
                lic_score = {'Academic 3.0': 39, 'AGPL 3.0': 91, 'Apache 1.0': 39, 'Apache 1.1': 39, 'Apache 2.0': 39, 'Apple 2.0': 52, 'Artistic 2.0': 65, 'Attribution Assurance': 39, 'Beerware': 39, 'Boost': 39, 'Bouncy Castle': 39, 'BSD 2': 39, 'BSD 3': 39, 'BSD 4': 39, 'CC BY 1.0': 39, 'CC BY SA 4.0': 39, 'CDDL 1.0': 52, 'CDDL 1.1': 52, 'CNRI Jython': 39, 'Common Public 1.0': 52, 'Computer Associates': 65, 'Eclipse 1.0': 65, 'Eclipse 2.0': 65, 'EDL 1.0': 39, 'Educational 2.0': 39, 'Eiffel Forum 2.0': 39, 'Entessa': 39, 'EU DataGrid': 39, 'Frameworx 1.0': 78, 'Golang BSD + Patents': 39, 'GPL 1.0': 78, 'GPL 2.0': 78, 'GPL 2.0 Classpath': 65, 'GPL 3.0': 78, 'Historical Permission': 39, 'IBM': 39, 'Illinois/NCSA': 39, 'ISC': 39, 'LGPL 2.0': 65, 'LGPL 2.1': 65, 'LGPL 3.0': 65, 'Lucent 1.02': 39, 'Microsoft Public': 65, 'Microsoft Reciprocal': 52, 'MIT': 39, 'Mozilla 1.0': 65, 'Mozilla 1.1': 65, 'Mozilla 2.0': 65, 'NUnit': 39, 'Open LDAP 2.4': 39, 'OpenSSL': 39, 'PostgreSQL': 39, 'Public Domain': 13, 'Python 2.0': 39, 'Ruby': 78, 'SIL Open Font 1.1': 39, 'Unlicense': 13, 'X.Net': 39, 'Zlib': 39}

                # Default dictionary to copyright score with associated risk
                score_labels = {13:"LOW",26:"LOW",39:"LOW",52:"LOW",65:"MEDIUM",78:"HIGH",91:"HIGH"}

                for key,value in lic_dic.items():
                        first_entry = key
                        if key in lic_score:
                        #       if lic_score[key] in score_labels:
                                second_entry = score_labels[lic_score[key]]
                        else:
                                second_entry = 'Not Known'
                        third_entry = value
                        thirddataframe.loc[len(thirddataframe.index)] = [first_entry, second_entry,third_entry]

                ws = wb.create_sheet("License risks",index=0)

                # Adding Project, repository name and build-id of the project
                ws["A1"] = f"Project Name : {pname}"
                ws["A1"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A1"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A1"].font = Font(bold=True)
                ws["A1"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                
                ws["A3"] = f"Repository Name : {rname}"
                ws["A3"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A3"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A3"].font = Font(bold=True)
                ws["A3"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                
                ws["A5"] = f"Build ID : {args[3]}\n (Click here for Mend Dashboard)"
                ws["A5"].style = "Hyperlink"
                ws["A5"].hyperlink = f'{args[4]}'
                ws["A5"].alignment = Alignment(horizontal='center',vertical='center',wrap_text=True)
                ws["A5"].fill = PatternFill(fill_type='solid',fgColor='FD923F')
                ws["A5"].font = Font(bold=True,color='000000')
                ws["A5"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws.merge_cells(start_row = 1,start_column=1,end_row = 2,end_column=3)
                ws.merge_cells(start_row = 3,start_column=1,end_row = 4,end_column=3)
                ws.merge_cells(start_row = 5,start_column=1,end_row = 6,end_column=3)
                ws.freeze_panes = ws["D7"]

                # Creating the main heading for the table

                ws["A8"] = "License Risks"
                ws["A8"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A8"].fill = PatternFill(fill_type='solid',fgColor='FFFF00')
                ws["A8"].font = Font(bold=True)
                # ws["B1"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                # ws["C1"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                ws["A8"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))
                ws.merge_cells(start_row = 8,start_column=1,end_row = 9,end_column=3)


                # Creating the table headers (for 'Library' and 'Licenses') in the 'Inventory' sheet
                ws["A10"] = 'License'
                ws["A10"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["A10"].alignment = Alignment(horizontal='center',vertical='center')
                ws["A10"].font = Font(bold=True)
                ws.column_dimensions['A'].width = 35
                ws["A10"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["B10"] = 'Risk'
                ws["B10"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["B10"].alignment = Alignment(horizontal='center',vertical='center')
                ws["B10"].font = Font(bold=True)
                ws.column_dimensions['B'].width = 20
                ws["B10"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws["C10"] = 'Occurrences'
                ws["C10"].fill = PatternFill(fill_type='solid',fgColor='A4DBE8')
                ws["C10"].alignment = Alignment(horizontal='center',vertical='center')
                ws["C10"].font = Font(bold=True)
                ws.column_dimensions['C'].width = 20
                ws["C10"].border = Border(top=Side("medium"),bottom=Side("medium"),left=Side("medium"),right=Side("medium"))

                ws.merge_cells(start_row = 10,start_column=1,end_row = 11,end_column=1)
                ws.merge_cells(start_row = 10,start_column=2,end_row = 11,end_column=2)
                ws.merge_cells(start_row = 10,start_column=3,end_row = 11,end_column=3) 

                # Write the DataFrame data to the specified location in the worksheet
                for idx, row in thirddataframe.iterrows():
                        for col_idx, value in enumerate(row, start=1):
                                ws.cell(row=12 + idx, column=col_idx, value=value)
                                ws.cell(row=12 + idx, column=col_idx).border = thin_border
                                ws.cell(row=12 + idx, column=col_idx).alignment = Alignment(horizontal='center',vertical='center') 

                wb.save(f"{args[2]}_mend_report.xlsx")

        except Exception as e:
                print(f"\nGetting the error {str(e)}")

if __name__ == "__main__":

        url_name = f"{sys.argv[5]}&view=whitesource.whiteSource-bolt-v2.build-tab.wss"
        start_creating_workbook(sys.argv[1],sys.argv[2],sys.argv[3],sys.argv[4],url_name)